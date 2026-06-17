#!/usr/bin/env python3
"""HIRA 마스터파일 → PostgreSQL 적재 스크립트"""
from __future__ import annotations
import sys
import logging
from datetime import datetime, date
from pathlib import Path
import psycopg2
import psycopg2.extras
import pyxlsb
import openpyxl

INCOMING = Path(__file__).parent.parent / "release_files" / "hira_incoming"
DB_CONFIG = dict(host="localhost", port=5432, dbname="term", user="postgres",
                 password="julab123!", options="-c search_path=term")

logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s [%(levelname)s] %(message)s",
                    handlers=[logging.StreamHandler(sys.stdout)])
log = logging.getLogger(__name__)


def to_date(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, (date, datetime)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    if s in ("", "9999-12-31", "99991231"):
        return None
    for fmt in ("%Y-%m-%d", "%Y%m%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def to_num(v) -> float | None:
    try:
        return float(v) if v not in (None, "", "None") else None
    except (TypeError, ValueError):
        return None


# ─── 행위 ──────────────────────────────────────────────────────────────────────
XLSB = INCOMING / "★수가반영내역(26.6.1.시행)_전체판.xlsb"

SHEET_MAP = {
    "의치과_급여_전체":     ("의치과", "급여"),
    "의치과_비급여_전체":   ("의치과", "비급여"),
    "의치과_100대100_전체": ("의치과", "100대100"),
    "한방_급여_전체":       ("한방",   "급여"),
    "한방_비급여_전체":     ("한방",   "비급여"),
    "한방_100대100_전체":   ("한방",   "100대100"),
    "약국_급여_전체":       ("약국",   "급여"),
}


def load_행위(conn):
    log.info("행위 적재 시작: %s", XLSB.name)
    rows_buf: list[tuple] = []
    total = 0

    with conn.cursor() as cur:
        cur.execute("TRUNCATE hira_행위_code")
    conn.commit()

    with pyxlsb.open_workbook(str(XLSB)) as wb:
        for sheet_name, (진료구분, 급여구분) in SHEET_MAP.items():
            시트구분 = f"{진료구분}_{급여구분}"
            cnt = 0
            with wb.get_sheet(sheet_name) as ws:
                rows = ws.rows()
                headers = [c.v for c in next(rows)]
                h = {v: i for i, v in enumerate(headers) if v}

                for row in rows:
                    v = [c.v for c in row]
                    if not v:
                        continue
                    수가코드 = str(v[0]).strip() if v[0] else None
                    if not 수가코드:
                        continue

                    def gv(key, default=None):
                        idx = h.get(key)
                        if idx is None:
                            return default
                        return v[idx] if idx < len(v) else default

                    rows_buf.append((
                        수가코드,
                        to_date(gv('적용일자')),
                        str(gv('분류번호') or ""),
                        str(gv('한글명') or ""),
                        str(gv('영문명') or gv('산정명칭') or ""),
                        str(gv('1-2 구분') or gv('1_2구분') or ""),
                        str(gv('수술여부') or gv('시술구분') or ""),
                        to_num(gv('의원단가') or gv('한방병의원단가') or gv('단가')),
                        to_num(gv('병원급이상단가') or gv('병원단가')),
                        to_num(gv('상대가치점수')),
                        시트구분,
                        str(gv('장구분') or ""),
                        str(gv('절구분') or ""),
                        str(gv('세분류') or ""),
                        str(gv('산정명칭') or gv('한글명') or ""),
                    ))
                    cnt += 1

                    if len(rows_buf) >= 5000:
                        _flush_행위(conn, rows_buf)
                        rows_buf.clear()

            total += cnt
            log.info("  %s: %d행", sheet_name, cnt)

    if rows_buf:
        _flush_행위(conn, rows_buf)
    conn.commit()
    log.info("행위 적재 완료: 총 %d행", total)


def _flush_행위(conn, rows_buf):
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO hira_행위_code
               (수가코드,적용일자,분류번호,한글명,영문명,구분,수술여부,
                의원단가,병원단가,상대가치점수,시트구분,장구분,절구분,세분류,산정명칭)
               VALUES %s
               ON CONFLICT (수가코드) DO UPDATE SET
                 적용일자=EXCLUDED.적용일자, 한글명=EXCLUDED.한글명,
                 영문명=EXCLUDED.영문명, 의원단가=EXCLUDED.의원단가,
                 병원단가=EXCLUDED.병원단가, 상대가치점수=EXCLUDED.상대가치점수""",
            rows_buf, page_size=2000
        )


# ─── 약제 ──────────────────────────────────────────────────────────────────────
DRUG_XLSX = INCOMING / "230501_260501_적용약가파일_4.30.수정 1부.xlsx"


def load_약제(conn):
    log.info("약제 적재 시작: %s", DRUG_XLSX.name)

    with conn.cursor() as cur:
        cur.execute("TRUNCATE hira_약제_code")
    conn.commit()

    wb = openpyxl.load_workbook(str(DRUG_XLSX), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = next(rows_iter)
    h = {v: i for i, v in enumerate(headers) if v}

    buf: list[tuple] = []
    total = 0
    for row in rows_iter:
        코드 = str(row[h['제품코드']]).strip() if row[h['제품코드']] else None
        dt = to_date(row[h['적용시작일자']])
        if not 코드 or dt is None:
            continue
        buf.append((
            코드, dt,
            str(row[h.get('급여기준', 2)] or ""),
            to_num(row[h.get('상한가', 3)]),
            str(row[h.get('투여경로', 5)] or ""),
            str(row[h.get('제품명', 6)] or ""),
            str(row[h.get('규격', 7)] or ""),
            str(row[h.get('단위', 8)] or ""),
            str(row[h.get('업체명', 9)] or ""),
            str(row[h.get('분류번호', 10)] or ""),
            str(row[h.get('주성분코드', 11)] or ""),
            str(row[h.get('전문/일반', 12)] or ""),
        ))
        total += 1
        if len(buf) >= 5000:
            _flush_약제(conn, buf)
            buf.clear()

    if buf:
        _flush_약제(conn, buf)
    conn.commit()
    wb.close()
    log.info("약제 적재 완료: 총 %d행", total)


def _flush_약제(conn, buf):
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO hira_약제_code
               (제품코드,적용시작일자,급여기준,상한가,투여경로,제품명,규격,단위,업체명,분류번호,주성분코드,전문일반)
               VALUES %s
               ON CONFLICT (제품코드,적용시작일자) DO UPDATE SET
                 상한가=EXCLUDED.상한가, 제품명=EXCLUDED.제품명""",
            buf, page_size=2000
        )


# ─── 치료재료 ──────────────────────────────────────────────────────────────────
SUPPLY_XLSX = INCOMING / "★2026.6.1._적용_(인체조직포함)_파일(급여).xlsx"
SUPPLY_XLSX2 = INCOMING / "★2026.6.1._적용_(인체조직포함)_파일(비급여).xlsx"


def load_치료재료(conn):
    log.info("치료재료 적재 시작")

    with conn.cursor() as cur:
        cur.execute("TRUNCATE hira_치료재료_code")
    conn.commit()

    total = 0
    for fpath, 급여구분 in [
        (SUPPLY_XLSX,  "급여"),
        (SUPPLY_XLSX2, "비급여"),
    ]:
        wb = openpyxl.load_workbook(str(fpath), read_only=True)
        sheet_name = "급여_품목(인체조직포함)" if 급여구분 == "급여" else "비급여_품목(인체조직포함)"
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers = next(rows_iter)
        h = {v: i for i, v in enumerate(headers) if v}

        buf: list[tuple] = []
        cnt = 0
        for row in rows_iter:
            코드 = str(row[h['코드']]).strip() if row[h['코드']] else None
            if not 코드:
                continue
            buf.append((
                코드,
                to_date(row[h.get('최초등재일', 1)]),
                to_date(row[h.get('적용일자', 2)]),
                to_date(row[h.get('종료일자', 3)]),
                str(row[h.get('중분류', 4)] or ""),
                str(row[h.get('중분류코드', 5)] or ""),
                str(row[h.get('품명', 6)] or ""),
                str(row[h.get('규격', 7)] or ""),
                str(row[h.get('단위', 8)] or ""),
                to_num(row[h.get('상한금액', 9)]),
                str(row[h.get('제조회사', 10)] or ""),
                str(row[h.get('재질', 11)] or ""),
                급여구분,
            ))
            cnt += 1
            if len(buf) >= 3000:
                _flush_치료재료(conn, buf)
                buf.clear()

        if buf:
            _flush_치료재료(conn, buf)
        conn.commit()
        wb.close()
        total += cnt
        log.info("  %s %s: %d행", fpath.name, 급여구분, cnt)

    log.info("치료재료 적재 완료: 총 %d행", total)


def _flush_치료재료(conn, buf):
    with conn.cursor() as cur:
        psycopg2.extras.execute_values(
            cur,
            """INSERT INTO hira_치료재료_code
               (코드,최초등재일,적용일자,종료일자,중분류,중분류코드,품명,규격,단위,상한금액,제조회사,재질,급여구분)
               VALUES %s
               ON CONFLICT (코드) DO UPDATE SET
                 적용일자=EXCLUDED.적용일자, 품명=EXCLUDED.품명,
                 상한금액=EXCLUDED.상한금액, 급여구분=EXCLUDED.급여구분""",
            buf, page_size=2000
        )


# ─── main ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        load_행위(conn)
        load_약제(conn)
        load_치료재료(conn)
    finally:
        conn.close()
    log.info("=== 전체 적재 완료 ===")
