import openpyxl, csv, os, glob

files = glob.glob('/home/infoclinic/services/term/release_files/KCD-9/*.xlsx')
if not files:
    raise FileNotFoundError('KCD-9 xlsx not found')
EXCEL = files[0]
print(f'Using: {EXCEL}')
OUT_MAIN  = '/home/infoclinic/services/term/icd10/KCD-9/kcd9_main.tsv'
OUT_MORPH = '/home/infoclinic/services/term/icd10/KCD-9/kcd9_morph.tsv'

wb = openpyxl.load_workbook(EXCEL, read_only=True, data_only=True)

ws2 = wb['KCD-9 DB Masterfile']
seen = set()
rows_main = []
for row in ws2.iter_rows(min_row=4, values_only=True):
    code = str(row[2]).strip() if row[2] is not None else ''
    if not code or code == 'None': continue
    korean = str(row[5]).strip() if row[5] is not None else ''
    english = str(row[6]).strip() if row[6] is not None else ''
    is_ext = 1 if (row[8] is not None and str(row[8]).strip() not in ('', '0', 'None')) else 0
    if not korean and not english: continue
    if code not in seen:
        seen.add(code)
        rows_main.append((code, korean, english, is_ext))

with open(OUT_MAIN, 'w', encoding='utf-8', newline='') as f:
    w = csv.writer(f, delimiter='\t')
    w.writerow(['CODE','KOREAN_LABEL','ENGLISH_LABEL','IS_KCD_EXT'])
    w.writerows(rows_main)
print(f'kcd9_main.tsv: {len(rows_main)} rows')

ws3 = wb['4편 신생물의 형태분류']
rows_morph = []
for row in ws3.iter_rows(min_row=4, values_only=True):
    code = str(row[0]).strip() if row[0] is not None else ''
    if not code or code == 'None': continue
    korean = str(row[1]).strip() if row[1] is not None else ''
    english = str(row[2]).strip() if row[2] is not None else ''
    rows_morph.append((code, korean, english))

with open(OUT_MORPH, 'w', encoding='utf-8', newline='') as f:
    w = csv.writer(f, delimiter='\t')
    w.writerow(['CODE','KOREAN_LABEL','ENGLISH_LABEL'])
    w.writerows(rows_morph)
print(f'kcd9_morph.tsv: {len(rows_morph)} rows')
